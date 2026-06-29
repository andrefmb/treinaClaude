#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera o "data room" rico da Bastião Seguradora (caso fictício) para o treinamento:
- Bastiao_DataRoom.xlsx  — planilha multi-aba, complexa, COM erros plantados e lacunas
- documentos extras (.md) para cruzamento, comparação e tradução
- Bastiao_Apolice_Escaneada.html — "apólice escaneada" para o exercício de Visão
- Bastiao_Gabarito_Facilitador.md — gabarito com os erros plantados (só p/ facilitador)

Determinístico (seed fixo) — rode: python3 gen_dataroom.py
TODOS os dados são FICTÍCIOS, criados apenas para o treinamento.
"""
import os
import random
import datetime as _dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(ROOT, "arquivos")
os.makedirs(OUT, exist_ok=True)
RND = random.Random(20260423)

NAVY = "16365C"
GOLD = "B5872F"
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color=NAVY, bold=True, size=14)
NOTE_FONT = Font(color="7A6320", italic=True, size=10)
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SETORES = ["Construção", "Construção", "Construção", "Infraestrutura", "Energia",
           "Saneamento", "Transporte", "Indústria", "Imobiliário", "Serviços"]
REGioes = ["Sudeste", "Sudeste", "Sudeste", "Sul", "Nordeste", "Centro-Oeste", "Norte"]
RAMOS = ["Garantia de Execução", "Garantia de Execução", "Garantia Judicial",
         "Fiança Locatícia", "Riscos de Engenharia"]
RATINGS = ["AA", "A", "BBB", "BB+", "BB", "BB-", "B+", "B"]
CAUSAS = ["Atraso de obra", "Inadimplência contratual", "Abandono de obra",
          "Falha técnica", "Recuperação judicial do tomador", "Distrato"]

# Tomadores nomeados (alguns com perfil definido para os exercícios)
TOMADORES = [
    # nome, setor, regiao, rating, divida_ebitda, liquidez
    ("Construtora Pilar Engenharia S.A.", "Construção", "Sudeste", "BB-", 4.8, 0.9),
    ("Rodovida Concessões S.A.", "Infraestrutura", "Sudeste", "BBB", 2.9, 1.4),
    ("Aurora Energia e Obras Ltda.", "Energia", "Nordeste", "BB+", 3.6, 1.1),
    ("Saneabrasil Engenharia S.A.", "Saneamento", "Sul", "A", 2.1, 1.7),
    ("Metrópole Construções S.A.", "Construção", "Sudeste", "BB", 4.1, 1.0),
    ("Ferrovia Norte Sul Obras Ltda.", "Transporte", "Centro-Oeste", "BBB", 3.0, 1.3),
    ("Atlântico Infra S.A.", "Infraestrutura", "Nordeste", "BB", 3.9, 1.05),
    ("Solaris Engenharia Ltda.", "Energia", "Sul", "A", 1.8, 1.9),
    ("Horizonte Imobiliária S.A.", "Imobiliário", "Sudeste", "BB+", 3.4, 1.2),
    ("Vetor Indústria Pesada S.A.", "Indústria", "Sudeste", "BBB", 2.7, 1.5),
    ("Costa Verde Saneamento S.A.", "Saneamento", "Nordeste", "BB", 4.0, 1.0),
    ("Planalto Construtora Ltda.", "Construção", "Centro-Oeste", "B+", 5.2, 0.8),
    ("Litoral Engenharia S.A.", "Construção", "Nordeste", "BB-", 4.6, 0.95),
    ("Trilhos & Pontes S.A.", "Transporte", "Sul", "BBB", 3.1, 1.35),
    ("NovaUrbe Incorporações S.A.", "Imobiliário", "Sudeste", "BB", 3.8, 1.1),
    ("Bandeirantes Obras S.A.", "Construção", "Sudeste", "BB+", 3.3, 1.25),
    ("Caatinga Energia S.A.", "Energia", "Nordeste", "A", 2.0, 1.8),
    ("Pantanal Infra Ltda.", "Infraestrutura", "Centro-Oeste", "BB", 4.2, 0.98),
    ("Serra Azul Construções S.A.", "Construção", "Sul", "BBB", 2.8, 1.45),
    ("MetroRio Obras S.A.", "Transporte", "Sudeste", "BB-", 4.7, 0.92),
    ("Delta Saneamento Ltda.", "Saneamento", "Norte", "B+", 5.0, 0.85),
    ("Cidade Nova Engenharia S.A.", "Construção", "Sudeste", "BB", 3.7, 1.15),
    ("Amazônia Infra S.A.", "Infraestrutura", "Norte", "BB-", 4.5, 0.97),
    ("Forte Construções Ltda.", "Construção", "Nordeste", "BBB", 3.0, 1.3),
    ("Sol Nascente Energia S.A.", "Energia", "Sul", "A", 1.9, 1.85),
]

errors = []  # gabarito do facilitador


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def brl(v):
    return v  # mantemos número puro; formato aplicado na célula


# ===========================================================================
wb = Workbook()

# ---- Aba 0: Leia-me ----
ws = wb.active
ws.title = "Leia-me"
ws["A1"] = "Bastião Seguradora S.A. — Data Room (CASO FICTÍCIO)"
ws["A1"].font = TITLE_FONT
notes = [
    "",
    "Todos os dados são fictícios, criados apenas para o treinamento 'Dominando o Claude'.",
    "Seguradora de médio porte especializada em SEGURO GARANTIA. Prêmios ~R$ 1,8 bi/ano.",
    "",
    "Abas:",
    "  • Resumo da Carteira — série trimestral 1T24–4T25 (sinistralidade, índice combinado).",
    "  • Apólices — carteira detalhada (IS, taxa, prêmio, vigência, status).",
    "  • Tomadores — crédito, exposição e limite aprovado por tomador.",
    "  • Sinistros — avisos em aberto/encerrados, reservas e pagamentos.",
    "  • Triângulo de Desenvolvimento — sinistros pagos acumulados por ano de origem.",
    "  • Limites e Alçada — política de aprovação por valor, setor e tomador.",
    "  • Macro — Selic e IPCA mensais (para precificação).",
    "",
    "ATENÇÃO (para o exercício de verificação): este data room contém ALGUMAS",
    "inconsistências e lacunas propositais. Confie nos números só depois de conferir.",
]
for i, t in enumerate(notes, start=2):
    ws.cell(row=i, column=1, value=t)
    if "ATENÇÃO" in t or "inconsist" in t:
        ws.cell(row=i, column=1).font = NOTE_FONT
autosize(ws, [95])

# ---- Aba: Resumo da Carteira (trimestral) ----
ws = wb.create_sheet("Resumo da Carteira")
hdr = ["Trimestre", "Prêmio Emitido (R$)", "Sinistros Pagos (R$)", "Sinistralidade (%)",
       "Despesas (%)", "Índice Combinado (%)", "Resultado Subscrição (R$)"]
ws.append(hdr)
style_header(ws, 1, len(hdr))
quarters = ["1T24", "2T24", "3T24", "4T24", "1T25", "2T25", "3T25", "4T25"]
sin = [22, 26, 31, 35, 38, 40, 41, 42]          # sinistralidade crescente
desp = [36, 36, 37, 37, 37, 38, 38, 37]
premio0 = 380_000_000
for i, q in enumerate(quarters):
    premio = int(premio0 * (1 + 0.04 * i))
    sinistros = int(premio * sin[i] / 100)
    sinistralidade = round(sin[i], 1)
    ic = round(sin[i] + desp[i], 1)
    resultado = int(premio * (100 - ic) / 100)
    # ERRO PLANTADO 1: no 3T25 a sinistralidade exibida não bate com sinistros/prêmio
    if q == "3T25":
        sinistralidade = 36.0  # exibido errado (real ~41)
        errors.append("Resumo da Carteira / 3T25: Sinistralidade exibida 36,0% não bate "
                      "com Sinistros Pagos ÷ Prêmio Emitido (real ≈ 41%). Pedir ao Claude "
                      "recalcular a partir das colunas.")
    # LACUNA PLANTADA: índice combinado de 2T25 em branco
    row = [q, premio, sinistros, sinistralidade, desp[i], ic, resultado]
    if q == "2T25":
        row[5] = None
        errors.append("Resumo da Carteira / 2T25: Índice Combinado em branco. Claude deve "
                      "calcular Sinistralidade + Despesas (= 78%).")
    ws.append(row)
for r in range(2, 2 + len(quarters)):
    for c in (2, 3, 7):
        ws.cell(row=r, column=c).number_format = '#,##0'
autosize(ws, [10, 20, 20, 16, 14, 20, 22])

# ---- Aba: Tomadores ----
ws = wb.create_sheet("Tomadores")
hdr = ["Tomador", "Setor", "Região", "Rating", "Dívida/EBITDA", "Liquidez Corrente",
       "Exposição Total (R$)", "Limite Aprovado (R$)", "% do Limite", "Situação"]
ws.append(hdr)
style_header(ws, 1, len(hdr))
tom_exposure = {}
for idx, (nome, setor, reg, rating, de, liq) in enumerate(TOMADORES):
    limite = RND.choice([60, 80, 100, 120, 150, 180]) * 1_000_000
    if "Pilar" in nome:
        exposicao = 150_000_000          # antes da nova garantia
        limite = 180_000_000
    else:
        exposicao = int(limite * RND.uniform(0.35, 1.15))
    tom_exposure[nome] = (exposicao, limite)
    pct = round(exposicao / limite * 100, 1)
    situacao = "Acima do limite" if exposicao > limite else "Dentro do limite"
    # ERRO PLANTADO 2: um tomador acima do limite marcado como "Dentro do limite"
    if nome.startswith("Planalto"):
        exposicao = int(limite * 1.18)
        tom_exposure[nome] = (exposicao, limite)
        pct = round(exposicao / limite * 100, 1)
        situacao = "Dentro do limite"   # ERRADO de propósito
        errors.append("Tomadores / Planalto Construtora: Exposição (118%% do limite) está "
                      "marcada como 'Dentro do limite' — deveria ser 'Acima do limite'.")
    # LACUNA PLANTADA: um rating em branco
    if nome.startswith("Litoral"):
        rating = None
        errors.append("Tomadores / Litoral Engenharia: Rating em branco — exige due "
                      "diligence antes de decidir.")
    row = [nome, setor, reg, rating, de, liq, exposicao, limite, pct, situacao]
    ws.append(row)
for r in range(2, 2 + len(TOMADORES)):
    for c in (7, 8):
        ws.cell(row=r, column=c).number_format = '#,##0'
autosize(ws, [34, 16, 14, 8, 14, 16, 20, 20, 11, 18])

# ---- Aba: Apólices ----
ws = wb.create_sheet("Apólices")
hdr = ["Apólice", "Ramo", "Tomador", "Setor", "Região", "IS (R$)", "Taxa (%)",
       "Prêmio (R$)", "Início Vigência", "Fim Vigência", "Status"]
ws.append(hdr)
style_header(ws, 1, len(hdr))
n_apolices = 220
planted_premio = 0
planted_neg = 0
for i in range(n_apolices):
    nome, setor, reg, rating, de, liq = RND.choice(TOMADORES)
    ramo = RND.choice(RAMOS)
    IS = RND.choice([5, 8, 10, 12, 15, 20, 25, 30, 40, 60]) * 1_000_000
    taxa = round(RND.uniform(0.6, 2.8), 2)
    premio = round(IS * taxa / 100)
    y = RND.choice([2024, 2025])
    m = RND.randint(1, 12)
    inicio = _dt.date(y, m, RND.randint(1, 28))
    fim = inicio + _dt.timedelta(days=RND.choice([365, 540, 730]))
    status = RND.choices(["Vigente", "Encerrada", "Em análise"], weights=[7, 2, 1])[0]
    regiao = reg
    # variação de grafia plantada (sujeira de dados)
    if i in (17, 88, 142):
        regiao = reg.upper()
    # ERRO PLANTADO 3: alguns prêmios não batem com IS x taxa
    if i in (12, 73, 150, 199) and planted_premio < 4:
        premio = round(premio * RND.uniform(1.25, 1.6))
        planted_premio += 1
    # LACUNA PLANTADA: algumas taxas em branco
    taxa_cell = taxa
    if i in (44, 111, 188):
        taxa_cell = None
    # ERRO PLANTADO 4: um IS negativo (typo)
    if i == 60 and planted_neg == 0:
        IS = -IS
        planted_neg += 1
    ap_id = "AP-%05d" % (10001 + i)
    ws.append([ap_id, ramo, nome, setor, regiao, IS, taxa_cell, premio,
               inicio.isoformat(), fim.isoformat(), status])
errors.append("Apólices: ~4 prêmios não batem com IS × Taxa (linhas plantadas); "
              "3 taxas em branco; 1 IS negativo (typo, AP com IS < 0); e 3 regiões "
              "grafadas em CAIXA ALTA (sujeira). Bom para pedir uma 'auditoria de qualidade'.")
for r in range(2, 2 + n_apolices):
    for c in (6, 8):
        ws.cell(row=r, column=c).number_format = '#,##0'
autosize(ws, [12, 22, 34, 16, 12, 16, 9, 16, 15, 14, 12])

# ---- Aba: Sinistros ----
ws = wb.create_sheet("Sinistros")
hdr = ["Sinistro", "Apólice", "Tomador", "Data Aviso", "Causa", "Status",
       "Valor Estimado (R$)", "Reserva (R$)", "Valor Pago (R$)"]
ws.append(hdr)
style_header(ws, 1, len(hdr))
n_sin = 42
for i in range(n_sin):
    nome, setor, reg, rating, de, liq = RND.choice(TOMADORES)
    sid = "SIN-%04d" % (2001 + i)
    apid = "AP-%05d" % (10001 + RND.randint(0, n_apolices - 1))
    y = RND.choice([2024, 2025])
    data = _dt.date(y, RND.randint(1, 12), RND.randint(1, 28))
    causa = RND.choice(CAUSAS)
    status = RND.choices(["Aberto", "Encerrado"], weights=[6, 4])[0]
    estimado = RND.choice([2, 3, 5, 8, 12, 18, 25, 40]) * 1_000_000
    if status == "Aberto":
        reserva = int(estimado * RND.uniform(0.6, 1.0)); pago = 0
    else:
        reserva = 0; pago = int(estimado * RND.uniform(0.5, 1.1))
    data_cell = data.isoformat()
    reserva_cell = reserva
    # ERRO PLANTADO 5: uma reserva como texto "n/d"
    if i == 9:
        reserva_cell = "n/d"
        errors.append("Sinistros / SIN-2010: Reserva preenchida como texto 'n/d' — quebra "
                      "somas. Claude deve sinalizar e tratar.")
    # ERRO PLANTADO 6: uma data em formato diferente (br)
    if i == 21:
        data_cell = data.strftime("%d/%m/%Y")
        errors.append("Sinistros / SIN-2022: Data de aviso em formato dd/mm/aaaa "
                      "(as demais em aaaa-mm-dd) — inconsistência de formato.")
    ws.append([sid, apid, nome, data_cell, causa, status, estimado, reserva_cell, pago])
# ERRO PLANTADO 7: id duplicado
ws.append(["SIN-2001", "AP-10099", "Construtora Pilar Engenharia S.A.",
           "2025-09-15", "Atraso de obra", "Aberto", 30_000_000, 24_000_000, 0])
errors.append("Sinistros: o id SIN-2001 aparece DUPLICADO (duas linhas). Claude deve "
              "detectar a duplicidade ao agregar.")
for r in range(2, ws.max_row + 1):
    for c in (7, 8, 9):
        cell = ws.cell(row=r, column=c)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'
autosize(ws, [10, 12, 34, 14, 26, 12, 20, 18, 18])

# ---- Aba: Triângulo de Desenvolvimento ----
ws = wb.create_sheet("Triângulo de Desenvolvimento")
ws["A1"] = "Sinistros pagos acumulados (R$ mil) por ano de origem × período de desenvolvimento"
ws["A1"].font = NOTE_FONT
hdr = ["Ano de Origem", "Dev 1", "Dev 2", "Dev 3", "Dev 4", "Dev 5"]
ws.append([])  # linha 2 vazia
ws.append(hdr)
style_header(ws, 3, len(hdr))
base = {2021: 4200, 2022: 5100, 2023: 6800, 2024: 9200, 2025: 12500}
factors = [1.0, 1.55, 1.9, 2.1, 2.2]
for yr, b in base.items():
    row = [yr]
    n = 2026 - yr  # quantos períodos já se desenvolveram
    for d in range(5):
        if d < n:
            row.append(int(b * factors[d]))
        else:
            row.append(None)  # ainda não desenvolvido (triângulo)
    ws.append(row)
for r in range(4, 9):
    for c in range(2, 7):
        cell = ws.cell(row=r, column=c)
        if isinstance(cell.value, int):
            cell.number_format = '#,##0'
autosize(ws, [16, 12, 12, 12, 12, 12])

# ---- Aba: Limites e Alçada ----
ws = wb.create_sheet("Limites e Alçada")
ws["A1"] = "Política de alçada (CASO FICTÍCIO)"
ws["A1"].font = TITLE_FONT
ws.append([])
ws.append(["Faixa de valor da garantia (R$)", "Quem aprova"])
style_header(ws, 3, 2)
alcada = [["Até 5 milhões", "Analista de subscrição"],
          ["5 a 20 milhões", "Gerente de subscrição"],
          ["20 a 50 milhões", "Comitê de subscrição"],
          ["Acima de 50 milhões", "Comitê + Diretoria de Risco"]]
for r in alcada:
    ws.append(r)
ws.append([])
ws.append(["Limite por setor (R$)", "Limite agregado"])
style_header(ws, ws.max_row, 2)
setor_lim = [["Construção Civil", 600_000_000], ["Infraestrutura", 400_000_000],
             ["Energia", 300_000_000], ["Saneamento", 250_000_000],
             ["Demais setores", 200_000_000]]
for nome, v in setor_lim:
    ws.append([nome, v]); ws.cell(row=ws.max_row, column=2).number_format = '#,##0'
ws.append([])
ws.append(["Regra de concentração", "Teto"])
style_header(ws, ws.max_row, 2)
ws.append(["Exposição máxima por tomador", "R$ 180.000.000"])
ws.append(["Concentração máxima em um setor", "35% da carteira"])
ws.append(["Concentração máxima em uma região", "45% da carteira"])
autosize(ws, [36, 28])

# ---- Aba: Macro ----
ws = wb.create_sheet("Macro")
hdr = ["Mês", "Selic (% a.a.)", "IPCA (% no mês)", "IPCA 12m (%)"]
ws.append(hdr)
style_header(ws, 1, len(hdr))
selic = 10.5
ipca12 = 4.2
for k in range(24):
    y = 2024 + (k + 4) // 12
    m = (k + 4) % 12 + 1
    selic += RND.choice([-0.25, 0, 0, 0.25, 0.5])
    selic = round(min(15.0, max(9.0, selic)), 2)
    ipca_m = round(RND.uniform(0.15, 0.62), 2)
    ipca12 = round(min(6.5, max(3.2, ipca12 + RND.uniform(-0.2, 0.25))), 2)
    ws.append(["%04d-%02d" % (y, m), selic, ipca_m, ipca12])
autosize(ws, [10, 16, 18, 14])

wb.save(os.path.join(OUT, "Bastiao_DataRoom.xlsx"))
print("  Bastiao_DataRoom.xlsx —", len(wb.sheetnames), "abas")

# ===========================================================================
# Gabarito do facilitador
# ===========================================================================
gab = ["# Gabarito do facilitador — erros e lacunas plantados\n",
       "> CONFIDENCIAL — só para o facilitador. Não distribua aos participantes.",
       "> Estes itens existem para treinar a disciplina de verificação (Módulo de governança).\n",
       "O `Bastiao_DataRoom.xlsx` contém inconsistências propositais. No exercício de",
       "verificação, peça ao Claude para auditar a qualidade dos dados — ele deve achar:\n"]
for i, e in enumerate(errors, 1):
    gab.append("%d. %s" % (i, e.replace("%%", "%")))
gab += ["\n## Respostas esperadas (números do caso)",
        "- Sinistralidade real do 3T25 ≈ 41% (não 36%).",
        "- Índice Combinado do 2T25 = 40 + 38 = 78%.",
        "- Construtora Pilar: antes da nova garantia, exposição R$ 150 mi / limite R$ 180 mi (83%).",
        "  Com a nova garantia de R$ 60 mi → R$ 210 mi → ESTOURA o limite de R$ 180 mi.",
        "- Planalto Construtora está, de fato, acima do limite (≈118%).",
        "- Pedir sempre as premissas na precificação; tratar como direcional."]
with open(os.path.join(OUT, "Bastiao_Gabarito_Facilitador.md"), "w", encoding="utf-8") as f:
    f.write("\n".join(gab) + "\n")
print("  Bastiao_Gabarito_Facilitador.md —", len(errors), "itens plantados")
print("Data room (planilha + gabarito) gerado.")
