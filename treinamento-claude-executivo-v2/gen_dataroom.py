#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera o data room fictício da AURORA S.A. (V2 executivo) para o treinamento:
- Aurora_DataRoom.xlsx  — planilha multi-aba (financeiro, mercado, concorrentes) COM erros plantados
- Aurora_Release_Resultados.md  — release de resultados (o "relatório" para virar deck)
- Aurora_Mercado.md  — panorama de mercado (para análise de mercado)
- Aurora_Concorrentes.md  — perfis dos concorrentes (para análise crítica)
- Aurora_Gabarito_Facilitador.md  — erros plantados (só facilitador)
Determinístico. Rode: python3 gen_dataroom.py
TODOS os dados são FICTÍCIOS, criados apenas para o treinamento.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(ROOT, "arquivos")
os.makedirs(OUT, exist_ok=True)

NAVY = "16365C"
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color=NAVY, bold=True, size=14)
NOTE_FONT = Font(color="7A6320", italic=True, size=10)
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
errors = []


def style_header(ws, row, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.border = BORDER
        cell.alignment = Alignment(vertical="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---- Leia-me ----
ws = wb.active
ws.title = "Leia-me"
ws["A1"] = "Aurora S.A. — Data Room (CASO FICTÍCIO)"
ws["A1"].font = TITLE_FONT
notes = [
    "",
    "Todos os dados são fictícios, criados apenas para o treinamento 'Dominando o Claude'.",
    "Aurora S.A.: varejo de bens de consumo (eletrônicos, casa e estilo de vida) — lojas + e-commerce.",
    "Receita ~R$ 4,2 bi/ano. Tensão central: crescimento desacelera, margem sob pressão e um",
    "concorrente digital (MercadoVivo) ganha share rápido. Decisão em pauta: acelerar o digital?",
    "",
    "Abas: Resultados Trimestrais · DRE 2025 · Balanço 2025 · Mercado & Share ·",
    "Concorrentes · Vendas por Região/Canal.",
    "",
    "ATENÇÃO (exercício de verificação): este data room tem algumas inconsistências e",
    "lacunas propositais. Confie nos números só depois de conferir.",
]
for i, t in enumerate(notes, 2):
    ws.cell(row=i, column=1, value=t)
    if "ATENÇÃO" in t or "inconsist" in t:
        ws.cell(row=i, column=1).font = NOTE_FONT
autosize(ws, [98])

# ---- Resultados Trimestrais ----
ws = wb.create_sheet("Resultados Trimestrais")
hdr = ["Trimestre", "Receita (R$ mi)", "Crescimento a/a (%)", "Margem Bruta (%)",
       "EBITDA (R$ mi)", "Margem EBITDA (%)", "Lucro Líquido (R$ mi)"]
ws.append(hdr); style_header(ws, 1, len(hdr))
quarters = ["1T24", "2T24", "3T24", "4T24", "1T25", "2T25", "3T25", "4T25"]
receita = [920, 980, 1010, 1180, 940, 990, 1005, 1160]
cresc =   [12.0, 11.0, 9.5, 8.0, 2.2, 1.0, -0.5, -1.7]
mbruta =  [34.0, 33.5, 33.0, 32.5, 32.0, 31.5, 31.0, 30.5]
for i, q in enumerate(quarters):
    rec = receita[i]
    ebitda = round(rec * (mbruta[i] - 20) / 100, 1)   # margem EBITDA ~ bruta - opex
    meb = round(ebitda / rec * 100, 1)
    ll = round(ebitda * 0.45, 1)
    meb_cell = meb
    # ERRO 1: 3T25 margem EBITDA exibida não bate com EBITDA/Receita
    if q == "3T25":
        meb_cell = 15.0
        errors.append("Resultados Trimestrais / 3T25: Margem EBITDA exibida 15,0% nao bate "
                      "com EBITDA / Receita (real ~ " + str(meb) + "%). Pedir para recalcular.")
    row = [q, rec, cresc[i], mbruta[i], ebitda, meb_cell, ll]
    # LACUNA: crescimento a/a do 2T25 em branco
    if q == "2T25":
        row[2] = None
        errors.append("Resultados Trimestrais / 2T25: Crescimento a/a em branco — calcular "
                      "vs. 2T24 (990 vs 980 ≈ +1,0%).")
    ws.append(row)
for r in range(2, 2 + len(quarters)):
    for c in (2, 5, 7):
        ws.cell(row=r, column=c).number_format = '#,##0.0'
autosize(ws, [10, 16, 19, 16, 15, 17, 19])

# ---- DRE 2025 ----
ws = wb.create_sheet("DRE 2025")
ws.append(["Demonstração de Resultados — 2025 (R$ mi)", ""])
ws["A1"].font = TITLE_FONT
ws.append(["Linha", "Valor"]); style_header(ws, 2, 2)
dre = [
    ("Receita líquida", 4095), ("(-) CMV", -2810), ("Lucro bruto", 1285),
    ("(-) Despesas com vendas", -640), ("(-) Despesas administrativas", -300),
    ("EBITDA", 345), ("(-) Depreciação e amortização", -110), ("EBIT", 235),
    ("(+/-) Resultado financeiro", -85), ("(-) Impostos", -45),
    ("Lucro líquido", 105),
]
for nome, val in dre:
    ws.append([nome, val]); ws.cell(row=ws.max_row, column=2).number_format = '#,##0'
# ERRO 2: lucro bruto não bate (4095-2810=1285 ok; vamos plantar no EBITDA)
errors.append("DRE 2025: EBITDA exibido (345) não fecha com Lucro bruto − Despesas "
              "(1285 − 640 − 300 = 345 está correto) — porém EBIT (235) usa D&A 110; "
              "confira a cadeia. (Plantado: ver Margem EBITDA da aba trimestral.)")
autosize(ws, [34, 14])

# ---- Balanço 2025 ----
ws = wb.create_sheet("Balanço 2025")
ws.append(["Balanço Patrimonial — 2025 (R$ mi)", ""]); ws["A1"].font = TITLE_FONT
ws.append(["Conta", "Valor"]); style_header(ws, 2, 2)
bal = [
    ("Ativo circulante", 1850), ("Ativo não circulante", 2050), ("TOTAL DO ATIVO", 3900),
    ("Passivo circulante", 1450), ("Passivo não circulante", 1300),
    ("Patrimônio líquido", 1100), ("TOTAL PASSIVO + PL", 3850),
]
for nome, val in bal:
    ws.append([nome, val]); ws.cell(row=ws.max_row, column=2).number_format = '#,##0'
errors.append("Balanço 2025: TOTAL DO ATIVO (3.900) ≠ TOTAL PASSIVO + PL (3.850) — "
              "diferença de R$ 50 mi. Um balanço tem de fechar; pedir para localizar.")
autosize(ws, [30, 14])

# ---- Mercado & Share ----
ws = wb.create_sheet("Mercado & Share")
hdr = ["Segmento", "TAM (R$ bi)", "Cresc. mercado (%)", "Share Aurora (%)",
       "Share líder (%)", "Líder do segmento"]
ws.append(hdr); style_header(ws, 1, len(hdr))
seg = [
    ("Eletrônicos", 58, 6.5, 7.2, 19.0, "MercadoVivo"),
    ("Casa & Decoração", 41, 4.0, 11.5, 16.0, "CasaBoa"),
    ("Estilo de vida", 33, 8.5, 5.0, 14.0, "Lumière"),
    ("Mercado/Marketplace", 90, 14.0, 1.8, 28.0, "MercadoVivo"),
]
for r in seg:
    ws.append(list(r))
errors.append("Mercado & Share: no segmento Eletrônicos, Share Aurora (7,2%) + Share líder "
              "(19,0%) + outros players citados ultrapassam 100% ao somar com a aba "
              "Concorrentes — checar a base e o período de cada fonte.")
autosize(ws, [22, 13, 18, 16, 15, 18])

# ---- Concorrentes ----
ws = wb.create_sheet("Concorrentes")
hdr = ["Concorrente", "Receita (R$ bi)", "Crescimento a/a (%)", "Margem EBITDA (%)",
       "Canal principal", "Nota"]
ws.append(hdr); style_header(ws, 1, len(hdr))
comp = [
    ("Aurora (nós)", 4.1, -0.5, 8.4, "Lojas + e-commerce", "Líder histórico em casa; digital atrasado"),
    ("MercadoVivo", 12.0, 22.0, 5.5, "Marketplace digital", "Cresce rápido, queima margem, ganha share"),
    ("Lumière Varejo", 6.8, 7.0, 9.0, "Lojas premium", "Marca forte em estilo de vida"),
    ("PontoCerto", 5.2, 3.0, 7.5, "Lojas + e-commerce", "Concorrente direto, preço agressivo"),
    ("CasaBoa", 3.4, 5.0, 10.5, "Lojas de bairro", "Forte em casa & decoração"),
]
for r in comp:
    ws.append(list(r))
# LACUNA: margem de um concorrente em branco
ws.cell(row=4, column=4, value=None)
errors.append("Concorrentes / Lumière: Margem EBITDA em branco — buscar em fonte pública "
              "antes de comparar.")
autosize(ws, [18, 16, 18, 16, 19, 42])

# ---- Vendas por Região/Canal ----
ws = wb.create_sheet("Vendas por Região")
hdr = ["Região", "Canal", "Receita (R$ mi)", "% do total", "Cresc. a/a (%)"]
ws.append(hdr); style_header(ws, 1, len(hdr))
vendas = [
    ("Sudeste", "Lojas", 1180, 28.8, -3.0),
    ("Sudeste", "E-commerce", 540, 13.2, 18.0),
    ("Sul", "Lojas", 620, 15.1, -1.0),
    ("Sul", "E-commerce", 240, 5.9, 22.0),
    ("NORDESTE", "Lojas", 560, 13.7, 1.0),   # caixa alta plantada
    ("Nordeste", "E-commerce", 190, 4.6, 25.0),
    ("Centro-Oeste", "Lojas", 330, 8.0, 0.0),
    ("Norte", "Lojas", 235, 5.7, -2.0),
]
for r in vendas:
    ws.append(list(r))
errors.append("Vendas por Região: os '% do total' somam ~95% (não 100%) e 'NORDESTE' "
              "aparece em CAIXA ALTA (as demais não) — pedir auditoria de consistência.")
for r in range(2, 2 + len(vendas)):
    ws.cell(row=r, column=3).number_format = '#,##0'
autosize(ws, [14, 14, 16, 12, 14])

wb.save(os.path.join(OUT, "Aurora_DataRoom.xlsx"))
print("  Aurora_DataRoom.xlsx —", len(wb.sheetnames), "abas")

# ---- Gabarito ----
gab = ["# Gabarito do facilitador — Aurora (erros e lacunas plantados)\n",
       "> CONFIDENCIAL — só para o facilitador. Não distribua antes do exercício de verificação.\n",
       "O `Aurora_DataRoom.xlsx` tem inconsistências propositais. O Claude deve achar:\n"]
for i, e in enumerate(errors, 1):
    gab.append("%d. %s" % (i, e.replace("%%", "%")))
gab += ["\n## Respostas esperadas (números do caso)",
        "- Crescimento desacelerou e ficou negativo no fim de 2025 (de +12% para ~-1,7%).",
        "- Margem bruta em queda lenta (34% → 30,5%); margem EBITDA real do 3T25 ≈ valor calculado, não 15%.",
        "- Balanço não fecha por R$ 50 mi — erro a localizar.",
        "- MercadoVivo é a maior ameaça: cresce ~22% e ganha share, mas com margem baixa (queima caixa).",
        "- A grande questão estratégica: acelerar o digital/marketplace (onde a Aurora tem só 1,8% de share)."]
with open(os.path.join(OUT, "Aurora_Gabarito_Facilitador.md"), "w", encoding="utf-8") as f:
    f.write("\n".join(gab) + "\n")
print("  Aurora_Gabarito_Facilitador.md —", len(errors), "itens")
print("Data room Aurora gerado.")
